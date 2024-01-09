#!venv/bin/python3
# read contacts from xsl files
# write to contact manager software

import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os


def extractSheet(sheet):

  sheet_entries=[]

  for row in sheet.iter_rows(min_row=2):
    drow={}
    for cell in row:
      if sheet.cell(row=1, column=cell.column).value and cell.value:
        drow[sheet.cell(row=1, column=cell.column).value]=str(cell.value)
    if drow != {}:
      sheet_entries.append(drow)

  return sheet_entries

def extractWorkbook(wb):

  wb_entries=[]

  for sheet in wb:
    print("'" + sheet.title + "'")
    wb_entries.extend(extractSheet(sheet))

  return wb_entries


def getHeadersForWorkbook(wb):

  headers=[]

  for sheet in wb:
    for row in sheet.iter_rows(max_row=2):
      for cell in row:
        if sheet.cell(row=1, column=cell.column).value:
          headers.append(sheet.cell(row=1, column=cell.column).value)

  return headers

def mergeListsWithOrderUnique(a, b):
  union = []
  for x in a + b:
      if x not in union:
          union.append(x)

  return union



directory = os.fsencode("./files")
all_entries=[]
headers=[]

for file in os.listdir(directory):
  filename = os.fsdecode(file)
  if filename.find("xls") >= 0:
    wb = load_workbook(filename=(os.path.join(directory.decode("utf-8"), filename)))
    print(wb.path)
    headers = mergeListsWithOrderUnique(headers, getHeadersForWorkbook(wb))
    print(headers)
#    time.sleep(5)

    all_entries.extend(extractWorkbook(wb))

    print(json.dumps(all_entries, indent=2))

output_path="./output/allentries.xlsx"

output_wb = openpyxl.Workbook()

output_sheet = output_wb.create_sheet("All entries", 0)

for column_idx in range(1, len(headers)):
  output_sheet.cell(row=1, column=column_idx).value=headers[column_idx-1]

output_sheet.freeze_panes = "A2"

for index, entry in enumerate(all_entries):
  row_index = index + 2
  for key, value in entry.items():
    print(key + " : " + value)
    header_idx = headers.index(key)
    column_idx = header_idx + 1

    output_sheet.cell(row=row_index, column=column_idx).value=value
  

output_wb.save(output_path)
 








